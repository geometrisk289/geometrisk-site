# -*- coding: utf-8 -*-
"""
Geometrisk — Find Your Line: data build.

Regenerates the site's data file from the Excel master, and refreshes the
offline fallback embedded in the page.

Master of truth for day-to-day edits is data/industry_losses.json (edit it in
GitHub or Claude Code and commit — the live site fetches it). Use THIS script
only for bulk edits made in the Excel workbook.

    python scripts/build_data.py

Reads : data/Geometrisk_Industry_Risk_Lookup.xlsx   (Industries + Losses sheets)
Writes: data/industry_losses.json                   (what the page fetches)
        find-your-line.html                          (offline fallback refreshed)
"""
import json, re, sys, os
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "data", "Geometrisk_Industry_Risk_Lookup.xlsx")
JSON_OUT = os.path.join(ROOT, "data", "industry_losses.json")
HTML = os.path.join(ROOT, "find-your-line.html")

def load_from_excel():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    wi, wl = wb["Industries"], wb["Losses"]
    inds, order = {}, []
    for row in wi.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        _id, label, code, div, kw = (list(row) + [None]*5)[:5]
        inds[_id] = dict(id=_id, label=label, anzsic=code or "", division=div or "",
                         keywords=kw or "", small=[], large=[])
        order.append(_id)
    for row in wl.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        _id, _label, band, loss, cost, peril = (list(row) + [None]*6)[:6]
        if _id not in inds:
            print(f"  ! losses row points at unknown industry id: {_id}", file=sys.stderr)
            continue
        band = (band or "").strip().lower()
        if band not in ("small", "large"):
            print(f"  ! bad band '{band}' for {_id}", file=sys.stderr)
            continue
        inds[_id][band].append([loss, int(cost), (peril or "").strip()])
    return [inds[i] for i in order]

def pretty(data):
    # one industry per block, compact loss rows — pleasant to hand-edit in GitHub
    parts = []
    for d in data:
        small = ",\n      ".join(json.dumps(x, ensure_ascii=False) for x in d["small"])
        large = ",\n      ".join(json.dumps(x, ensure_ascii=False) for x in d["large"])
        parts.append(
            "  {\n"
            f'    "id": {json.dumps(d["id"])},\n'
            f'    "label": {json.dumps(d["label"], ensure_ascii=False)},\n'
            f'    "anzsic": {json.dumps(d["anzsic"], ensure_ascii=False)},\n'
            f'    "division": {json.dumps(d["division"], ensure_ascii=False)},\n'
            f'    "keywords": {json.dumps(d["keywords"], ensure_ascii=False)},\n'
            f'    "small": [\n      {small}\n    ],\n'
            f'    "large": [\n      {large}\n    ]\n'
            "  }"
        )
    return "[\n" + ",\n".join(parts) + "\n]\n"

def main():
    data = load_from_excel()
    js = pretty(data)
    with open(JSON_OUT, "w", encoding="utf-8") as f:
        f.write(js)
    # refresh offline fallback in the page
    if os.path.exists(HTML):
        html = open(HTML, encoding="utf-8").read()
        block = '<script id="fallback-data" type="application/json">' + json.dumps(data, ensure_ascii=False) + '</script>'
        html2 = re.sub(r'<script id="fallback-data"[^>]*>.*?</script>', block, html, flags=re.S)
        open(HTML, "w", encoding="utf-8").write(html2)
    rows = sum(len(d["small"]) + len(d["large"]) for d in data)
    print(f"OK — {len(data)} industries, {rows} loss rows -> {os.path.relpath(JSON_OUT, ROOT)}")

if __name__ == "__main__":
    main()
